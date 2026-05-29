"""检查 Excel 年级花名册中 B 列姓名的格式规范。

检查规则：
- 两字姓名（去空格后 len=2）：字符间必须恰好 2 个 ASCII 空格
- 三字及以上姓名：不应含空格（· 除外）
- 含 · 的姓名：· 前后不应有空格
"""

import sys
import os
import re
sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from utils import get_excel_path, normalize_name


def check_two_char_name(name, sheet_name, row_num):
    """检查两字姓名的空格格式。"""
    stripped = name.strip()
    chars_no_space = normalize_name(name)
    if len(chars_no_space) != 2:
        return []

    issues = []
    space_count = stripped.count(' ')
    # 两个字符中间应该恰好有 2 个空格
    if space_count != 2:
        issues.append(
            f"[{sheet_name}] 第{row_num}行: '{stripped}' "
            f"两字姓名应有2个空格，实际有{space_count}个空格"
        )
    # 检查空格位置：应该在两个字中间
    cleaned = stripped.replace(' ', '')
    if len(cleaned) == 2:
        expected = cleaned[0] + '  ' + cleaned[1]
        if stripped != expected:
            issues.append(
                f"[{sheet_name}] 第{row_num}行: '{stripped}' "
                f"格式不正确，期望格式: '{expected}'"
            )
    return issues


def check_normal_name(name, sheet_name, row_num):
    """检查三字及以上姓名不应有空格。"""
    stripped = name.strip()
    chars_no_space = normalize_name(name)
    if len(chars_no_space) <= 2:
        return []
    if '·' in chars_no_space:
        return []  # 含 · 的名字单独检查

    if ' ' in stripped:
        return [
            f"[{sheet_name}] 第{row_num}行: '{stripped}' "
            f"三字及以上姓名不应含空格"
        ]
    return []


def check_dot_name(name, sheet_name, row_num):
    """检查含 · 的姓名格式。"""
    stripped = name.strip()
    if '·' not in stripped:
        return []

    issues = []
    # · 前后不应有空格
    if ' ·' in stripped or '· ' in stripped:
        issues.append(
            f"[{sheet_name}] 第{row_num}行: '{stripped}' "
            f"间隔号(·)前后不应有空格"
        )
    return issues


def main():
    excel_path = get_excel_path()
    print(f"检查文件: {excel_path}")
    print()

    wb = openpyxl.load_workbook(excel_path)
    all_issues = []
    stats = {'total': 0, 'two_char': 0, 'normal': 0, 'dot': 0}

    for sn in wb.sheetnames:
        ws = wb[sn]
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True),
            start=2
        ):
            name = row[0]
            if not name or not str(name).strip():
                continue
            name_str = str(name).strip()
            stats['total'] += 1

            clean = normalize_name(name_str)
            if '·' in name_str:
                stats['dot'] += 1
                all_issues.extend(check_dot_name(name_str, sn, row_idx))
            elif len(clean) == 2:
                stats['two_char'] += 1
                all_issues.extend(check_two_char_name(name_str, sn, row_idx))
            else:
                stats['normal'] += 1
                all_issues.extend(check_normal_name(name_str, sn, row_idx))

    wb.close()

    # 输出统计
    print("=" * 60)
    print("统计信息")
    print("=" * 60)
    print(f"  总姓名数:     {stats['total']}")
    print(f"  两字姓名:     {stats['two_char']}")
    print(f"  三字及以上:   {stats['normal']}")
    print(f"  含间隔号(·):  {stats['dot']}")
    print()

    # 输出问题
    if all_issues:
        print("=" * 60)
        print(f"发现 {len(all_issues)} 个问题:")
        print("=" * 60)
        for issue in all_issues:
            print(f"  {issue}")
    else:
        print("=" * 60)
        print("未发现任何格式问题，所有姓名格式规范！")
        print("=" * 60)

    return len(all_issues)


if __name__ == '__main__':
    sys.exit(main())

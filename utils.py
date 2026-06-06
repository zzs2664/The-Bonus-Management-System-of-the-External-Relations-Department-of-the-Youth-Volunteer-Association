"""青志协加分数据库 — 共享工具模块"""

import re
import os
import sys

# ============================================================
# 班级名
# ============================================================

def load_class_names(workbook):
    """从 Excel 读取所有 sheet 名作为合法班级名集合。"""
    return set(workbook.sheetnames)


def find_class_in_text(text, class_names):
    """在文本开头查找已知班级名，返回 (class_name, remaining_text) 或 (None, text)。"""
    for cn in sorted(class_names, key=len, reverse=True):
        if text.startswith(cn):
            remainder = text[len(cn):]
            return cn, remainder
    return None, text


# ============================================================
# 学生姓名
# ============================================================

def load_student_names(worksheet):
    """从 Sheet 的 B 列加载学生姓名。

    Returns:
        dict: {clean_name: original_name}
            clean_name   — 去空格后的名字，用于匹配
            original_name — Excel B 列原始值，用于写入
    """
    names = {}
    for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        name = row[0]
        if name and str(name).strip():
            name_str = str(name).strip()
            clean = normalize_name(name_str)
            if clean:
                names[clean] = name_str
    return names


def normalize_name(name):
    """去掉所有空格，用于名字比较。"""
    return re.sub(r'\s+', '', str(name))


# ============================================================
# 分数标题
# ============================================================

SCORE_HEADER_RE = re.compile(r'^\((\d+(?:\.\d+)?)分\)$')


def parse_score_header(text):
    """识别 '(1分)' '(0.9分)' 等，返回 float 分数或 None。"""
    m = SCORE_HEADER_RE.match(text)
    if m:
        return float(m.group(1))
    return None


# ============================================================
# 日期提取
# ============================================================

def extract_date_from_filename(filepath):
    """从文件名提取日期字符串。

    '3.29清瓶乐活动德育分统计.docx' → '3.29'
    '12.15某某活动德育分统计.docx'  → '12.15'
    """
    basename = os.path.basename(filepath)
    basename = os.path.splitext(basename)[0]
    m = re.match(r'^(\d{1,2}\.\d{1,2})', basename)
    if m:
        return m.group(1)
    return basename


# ============================================================
# 日期列查找（startswith 匹配）
# ============================================================

def find_date_columns(ws, date_str):
    """在 Sheet 第一行查找所有以 date_str 开头的日期列。

    '3.29' → 匹配 '3.29清瓶乐'、'3.29扫雪' 等

    Args:
        ws: openpyxl Worksheet 对象
        date_str: 日期前缀，如 '3.29'

    Returns:
        [(col_idx, header_value), ...] 按列号排序
    """
    matches = []
    for col_idx in range(3, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and str(cell_value).strip().startswith(date_str):
            matches.append((col_idx, str(cell_value).strip()))
    return matches


def extract_date_prefix(text):
    """从完整列头中提取纯日期前缀。

    '3.29清瓶乐' → '3.29'
    '12.15扫雪'  → '12.15'
    '3.29'       → '3.29'
    """
    m = re.match(r'^(\d{1,2}\.\d{1,2})', text)
    if m:
        return m.group(1)
    return text


def extract_column_header(filepath):
    """从文件名提取完整列头（日期+活动名）。

    '3.29清瓶乐活动德育分统计.docx' → '3.29清瓶乐'
    '12.15某某活动德育分统计.docx'  → '12.15某某'
    '3.29活动德育分统计.docx'       → '3.29'
    """
    basename = os.path.basename(filepath)
    basename = os.path.splitext(basename)[0]
    # 匹配 "数字.数字" 到 "活动" 之间的部分
    m = re.match(r'^(\d{1,2}\.\d{1,2}.*?)活动', basename)
    if m:
        return m.group(1).rstrip()
    return basename


def scan_excel_date_columns(excel_path, date_str):
    """扫描 Excel 所有 Sheet，返回匹配 date_str 的列头集合（去重+排序）。

    Returns:
        list of str — 如 ['3.29清瓶乐', '3.29扫雪']
    """
    if not excel_path or not os.path.exists(excel_path):
        return []
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    class_names = load_class_names(wb)
    headers = set()
    for cn in sorted(class_names):
        ws = wb[cn]
        cols = find_date_columns(ws, date_str)
        for _, header in cols:
            headers.add(header)
    wb.close()
    return sorted(headers)


# ============================================================
# 班级排序键
# ============================================================

CLASS_ORDER = {'地卓': 0, '地师': 1, '地理': 2, '地信': 3, '环境': 4}


def class_sort_key(class_name):
    """返回班级排序键：(年级, 专业优先级, 班级名)。

    年级优先：23 > 24 > 25
    同级按专业：地卓 > 地师 > 地理 > 地信 > 环境
    """
    match = re.match(r'^(\D+?)(?:本|类)?(\d{2})\d{2}$', class_name)
    if match:
        major_prefix = match.group(1)
        grade = int(match.group(2))
        major_order = CLASS_ORDER.get(major_prefix, 99)
        return (grade, major_order, class_name)
    return (99, 99, class_name)


# ============================================================
# 姓名长度排序键
# ============================================================

def name_length_key(name):
    """返回姓名长度排序键。

    优先级：3字 > 2字 > 4字 > 5字 > 6字 > 含点(·)

    返回 (priority, name) 其中 priority 越小越靠前。
    """
    clean = normalize_name(name)
    length = len(clean)
    has_dot = '·' in name

    if has_dot:
        return (6, name)
    if length == 3:
        return (0, name)
    if length == 2:
        return (1, name)
    if length == 4:
        return (2, name)
    if length == 5:
        return (3, name)
    if length >= 6:
        return (4, name)
    return (5, name)


# ============================================================
# 中文字符判断
# ============================================================

def is_chinese_char(ch):
    """判断是否为中文字符（CJK统一表意文字）。"""
    cp = ord(ch)
    return (0x4E00 <= cp <= 0x9FFF or
            0x3400 <= cp <= 0x4DBF or
            0x20000 <= cp <= 0x2A6DF or
            0xF900 <= cp <= 0xFAFF or
            0x2F800 <= cp <= 0x2FA1F)


def classify_char(ch):
    """将字符分类：'chinese', 'digit_punct'（数字/括号/点）, 'space', 'other'"""
    if ch in ('·',):
        return 'digit_punct'
    if is_chinese_char(ch):
        return 'chinese'
    if ch.isdigit():
        return 'digit_punct'
    if ch in ('(', ')', '.', '-', '+'):
        return 'digit_punct'
    if ch.isspace():
        return 'space'
    return 'other'


# ============================================================
# 路径工具
# ============================================================

def get_data_dir():
    """获取项目根目录。"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_output_dir():
    """获取输出文件夹路径（自动创建）。"""
    output_dir = os.path.join(get_data_dir(), '输出文件夹')
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def get_test_dir():
    """获取测试文件夹路径。"""
    return os.path.join(get_data_dir(), '测试文件夹')


def get_excel_path():
    """获取年级花名册数据表路径。"""
    return os.path.join(get_data_dir(), '年级花名册数据表.xlsx')

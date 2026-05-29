"""青志协加分数据库 — Word → Excel 录入脚本

用法:
    python import_scores.py "文件模板/3.29清瓶乐活动德育分统计.docx"

功能:
    1. 从 Word 加分名单中解析活动日期、分数、班级、姓名
    2. 将分数数据写入 Excel 年级花名册对应的班级 Sheet
    3. 支持交互式覆盖确认、错误日志输出
"""

import sys
import os
import re
import copy
from collections import defaultdict

sys.stdout.reconfigure(encoding='utf-8')
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from docx import Document

from utils import (
    load_class_names, load_student_names, normalize_name,
    parse_score_header, find_class_in_text, extract_date_from_filename,
    get_excel_path, get_output_dir
)


# ============================================================
# 姓名匹配核心算法
# ============================================================

def build_name_pattern(clean_name):
    """为姓名构建宽松匹配的正则表达式。

    在每个字符之间插入 \\s* 以允许任意空白（0、1、2个空格等）。
    使用 (?:^|\\s) 前缀和 (?=\\s|$) 后缀确保匹配完整姓名。

    Args:
        clean_name: 去空格后的干净姓名

    Returns:
        compiled regex pattern (capturing group 1 = the actual name)
    """
    chars = list(clean_name)
    inner = r'\s*'.join(re.escape(c) for c in chars)
    full_pattern = r'(?:^|\s)(' + inner + r')(?=\s|$)'
    return re.compile(full_pattern)


def extract_names_from_text(text, student_dict):
    """用字典匹配法从 Word 文本中提取已知学生姓名。

    使用 \\s* 宽松匹配：允许两字姓名中间的 0/1/2 个空格。
    长名优先匹配，避免短名误匹配到长名的子串。
    匹配到的部分会从文本中移除，避免重复匹配。

    Args:
        text: 去掉班级名前缀后的 Word 行文本
        student_dict: {clean_name: original_name} 从 Excel 加载

    Returns:
        (matched_names, unmatched_text)
        matched_names: list of original_name (Excel 原始格式)
        unmatched_text: 移除已匹配姓名后剩余的文本
    """
    matched = []
    # 按姓名长度降序排列：长名优先匹配
    sorted_names = sorted(student_dict.items(), key=lambda x: -len(x[0]))

    for clean_name, original_name in sorted_names:
        pattern = build_name_pattern(clean_name)
        m = pattern.search(text)
        if m:
            matched.append(original_name)
            # 从文本中移除已匹配部分（替换为空格，保持位置）
            # 使用 group(1) 的起止位置（不含前导空格）
            text = text[:m.start(1)] + ' ' + text[m.end(1):]

    # 清理剩余文本：去掉多余空格
    unmatched = re.sub(r'\s+', ' ', text).strip()
    return matched, unmatched


# ============================================================
# 兜底解析（当字典匹配不完全时的备用方案）
# ============================================================

def fallback_parse_names(text):
    """兜底解析：按任意空白 split，合并相邻单字为两字姓名。

    仅在字典匹配未能覆盖所有姓名时使用。
    结果写入警告日志。
    """
    if not text.strip():
        return []

    # 按任意连续空白分割
    tokens = re.split(r'\s+', text.strip())
    tokens = [t for t in tokens if t]

    names = []
    i = 0
    while i < len(tokens):
        token = tokens[i]
        # 单字 Token：与下一个 Token 合并为两字姓名
        if len(token) == 1 and i + 1 < len(tokens):
            next_token = tokens[i + 1]
            if len(next_token) == 1:
                names.append(token + next_token)
                i += 2
                continue
        names.append(token)
        i += 1

    return names


# ============================================================
# Word 文档解析
# ============================================================

def parse_word_doc(docx_path, class_names, student_cache):
    """解析 Word 加分名单文档。

    Args:
        docx_path: Word 文档路径
        class_names: 合法班级名集合 (set of str)
        student_cache: {class_name: {clean_name: original_name}}
                       预加载的所有班级学生姓名

    Returns:
        (records, warnings)
        records: [(class_name, original_name, score), ...]
        warnings: list of warning strings
    """
    doc = Document(docx_path)
    records = []
    warnings = []

    current_score = None
    current_class = None

    for i, p in enumerate(doc.paragraphs):
        text = p.text.strip()
        if not text:
            continue

        # 按换行符（Shift+Enter）拆分为多行分别处理
        sub_lines = text.split('\n')

        for line in sub_lines:
            line = line.strip()
            if not line:
                continue

            # 1. 检测分数标题
            score = parse_score_header(line)
            if score is not None:
                current_score = score
                current_class = None
                continue

            # 2. 检测班级行
            class_name, remainder = find_class_in_text(line, class_names)

            if class_name is not None:
                current_class = class_name

                # 检查班级是否有对应 Sheet
                if class_name not in student_cache:
                    warnings.append(
                        f"段落[{i}]: 班级 '{class_name}' 在 Excel 中无对应 Sheet，跳过"
                    )
                    current_class = None
                    continue

                # 用字典匹配提取姓名
                matched_names, unmatched = extract_names_from_text(
                    remainder, student_cache[class_name]
                )

                for name in matched_names:
                    records.append((class_name, name, current_score))

                # 兜底解析未匹配部分
                if unmatched:
                    fallback_names = fallback_parse_names(unmatched)
                    for fn in fallback_names:
                        warnings.append(
                            f"段落[{i}]: 班级 '{class_name}' 中姓名 '{fn}' "
                            f"未在 Excel 中找到，使用兜底解析"
                        )
                        records.append((class_name, fn, current_score))

            else:
                # 3. 续行（无班级名、无分数标题）→ 属于上一个班级
                if current_class is None:
                    warnings.append(
                        f"段落[{i}]: 续行但无当前班级上下文，跳过: '{line[:50]}...'"
                    )
                    continue

                if current_class not in student_cache:
                    continue

                matched_names, unmatched = extract_names_from_text(
                    line, student_cache[current_class]
                )

                for name in matched_names:
                    records.append((current_class, name, current_score))

                if unmatched:
                    fallback_names = fallback_parse_names(unmatched)
                    for fn in fallback_names:
                        warnings.append(
                            f"段落[{i}]: 续行中班级 '{current_class}' 的姓名 '{fn}' "
                            f"未在 Excel 中找到，使用兜底解析"
                        )
                        records.append((current_class, fn, current_score))

    return records, warnings


# ============================================================
# Excel 写入
# ============================================================

def find_or_create_date_column(ws, date_str):
    """在 Sheet 的第一行查找日期列，若不存在则在最后一列右侧创建。"""
    # 扫描第一行
    for col_idx in range(3, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value and str(cell_value).strip() == date_str:
            return col_idx

    # 不存在，新建列
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=date_str)
    return new_col


def write_scores_to_excel(excel_path, records, date_str):
    """将分数记录写入 Excel。

    Args:
        excel_path: Excel 文件路径
        records: [(class_name, original_name, score), ...]
        date_str: 日期列头（如 '3.29'）

    Returns:
        (output_path, missing_students, total_written)
    """
    wb = openpyxl.load_workbook(excel_path)
    class_names = load_class_names(wb)

    missing_students = []
    total_written = 0

    # 按班级分组
    by_class = defaultdict(list)
    for class_name, name, score in records:
        by_class[class_name].append((name, score))

    for class_name, entries in by_class.items():
        if class_name not in class_names:
            # 班级在 Excel 中无对应 Sheet（已在解析阶段警告过）
            continue

        ws = wb[class_name]

        # 构建 B 列姓名索引: {normalized_name: row_number}
        name_to_row = {}
        for row_idx in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=2).value
            if cell_value and str(cell_value).strip():
                clean = normalize_name(str(cell_value))
                name_to_row[clean] = row_idx

        # 找到或创建日期列（已有列则清空重写）
        col = None
        for col_idx in range(3, ws.max_column + 1):
            if str(ws.cell(row=1, column=col_idx).value or '').strip() == date_str:
                col = col_idx
                break
        if col is None:
            col = find_or_create_date_column(ws, date_str)
        else:
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=col, value=None)

        # 写入分数
        for name, score in entries:
            clean = normalize_name(name)
            if clean in name_to_row:
                row = name_to_row[clean]
                ws.cell(row=row, column=col, value=score)
                total_written += 1
            else:
                missing_students.append((class_name, name, score))

    # 保存
    output_filename = '年级花名册数据表_updated.xlsx'
    output_path = os.path.join(get_output_dir(), output_filename)
    wb.save(output_path)
    wb.close()

    return output_path, missing_students, total_written


# ============================================================
# 日志
# ============================================================

def write_error_log(output_dir, missing_students, warnings):
    """将未匹配学生和警告写入 error.log。"""
    log_path = os.path.join(output_dir, 'error.log')
    with open(log_path, 'w', encoding='utf-8') as f:
        f.write("=" * 60 + "\n")
        f.write("导入错误日志\n")
        f.write("=" * 60 + "\n\n")

        if missing_students:
            f.write(f"未在 Excel 中找到的学生 ({len(missing_students)} 人):\n")
            f.write("-" * 40 + "\n")
            for class_name, name, score in missing_students:
                f.write(f"  班级: {class_name}, 姓名: {name}, 分数: {score}\n")
            f.write("\n")

        if warnings:
            f.write(f"警告信息 ({len(warnings)} 条):\n")
            f.write("-" * 40 + "\n")
            for w in warnings:
                f.write(f"  {w}\n")

    return log_path


# ============================================================
# 主流程
# ============================================================

def main():
    if len(sys.argv) < 2:
        print("用法: python import_scores.py <path_to_docx> [--excel <path>]")
        print("示例: python import_scores.py \"文件模板/3.29清瓶乐活动德育分统计.docx\"")
        print("      python import_scores.py \"文件模板/3.29清瓶乐活动德育分统计.docx\" --excel \"输出文件夹/年级花名册数据表_updated.xlsx\"")
        sys.exit(1)

    docx_path = sys.argv[1]
    if not os.path.exists(docx_path):
        print(f"错误: 文件不存在: {docx_path}")
        sys.exit(1)

    # 解析 --excel 选项
    excel_path = None
    if '--excel' in sys.argv:
        idx = sys.argv.index('--excel')
        if idx + 1 < len(sys.argv):
            excel_path = sys.argv[idx + 1]

    print("=" * 60)
    print("青志协加分数据库 — 录入脚本 (Word → Excel)")
    print("=" * 60)
    print()

    # 1. 提取日期
    date_str = extract_date_from_filename(docx_path)
    print(f"活动日期: {date_str}")

    # 2. 加载 Excel 和学生姓名缓存
    if excel_path is None:
        updated_path = os.path.join(get_output_dir(), '年级花名册数据表_updated.xlsx')
        if os.path.exists(updated_path):
            excel_path = updated_path
        else:
            excel_path = get_excel_path()
    print(f"数据库:   {os.path.basename(excel_path)}")

    wb = openpyxl.load_workbook(excel_path)
    class_names = load_class_names(wb)

    student_cache = {}
    for cn in class_names:
        ws = wb[cn]
        student_cache[cn] = load_student_names(ws)
    wb.close()

    print(f"班级数:   {len(class_names)}")
    total_students = sum(len(v) for v in student_cache.values())
    print(f"学生总数: {total_students}")
    print()

    # 3. 解析 Word 文档
    print(f"解析 Word: {os.path.basename(docx_path)}")
    records, warnings = parse_word_doc(docx_path, class_names, student_cache)

    print(f"提取记录: {len(records)} 条")
    if warnings:
        print(f"警告数量: {len(warnings)} 条")
    print()

    # 汇总分数分布
    score_dist = defaultdict(int)
    for _, _, score in records:
        score_dist[score] += 1
    print("分数分布:")
    for s in sorted(score_dist.keys(), reverse=True):
        print(f"  {s}分: {score_dist[s]} 人")
    print()

    # 4. 写入 Excel
    print("写入 Excel...")
    output_path, missing_students, total_written = write_scores_to_excel(
        excel_path, records, date_str
    )

    if output_path is None:
        return  # 用户取消

    print(f"成功写入: {total_written} 条")
    print(f"输出文件: {output_path}")

    # 5. 错误日志
    if missing_students or warnings:
        log_path = write_error_log(get_output_dir(), missing_students, warnings)
        print(f"错误日志: {log_path}")

        if missing_students:
            print(f"\n未匹配学生 ({len(missing_students)} 人):")
            for class_name, name, score in missing_students:
                print(f"  [{class_name}] {name} ({score}分)")

    print()
    print("导入完成！")


if __name__ == '__main__':
    main()
